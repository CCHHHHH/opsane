"""Read-only attachment parsing.

The parser never executes uploaded files and never extracts archive members to disk.
Only bounded text and archive metadata are retained for previews and LLM context.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from functools import lru_cache
from hashlib import sha256
from html.parser import HTMLParser
import os
from pathlib import Path
import re
import signal
import shutil
import subprocess
import sys
import tarfile
import tempfile
from xml.etree import ElementTree
import zipfile


MAX_EXTRACTED_CHARS = 400_000
MAX_TEXT_BYTES = 4 * 1024 * 1024
MAX_XML_MEMBER_BYTES = 12 * 1024 * 1024
MAX_OFFICE_ARCHIVE_ENTRIES = 5_000
MAX_OFFICE_TOTAL_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 600
MAX_OCR_IMAGE_BYTES = 30 * 1024 * 1024
OCR_TIMEOUT_SECONDS = 45
MAX_LEGACY_OFFICE_BYTES = 100 * 1024 * 1024
MAX_OFFICE_PREVIEW_BYTES = 100 * 1024 * 1024
OFFICE_CONVERSION_TIMEOUT_SECONDS = 90

TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".out", ".conf", ".cfg",
    ".ini", ".properties", ".env", ".yaml", ".yml", ".json", ".jsonl",
    ".xml", ".html", ".htm", ".csv", ".tsv", ".sql", ".sh", ".bash",
    ".zsh", ".py", ".js", ".ts", ".tsx", ".jsx", ".vue", ".java",
    ".go", ".rs", ".c", ".h", ".cpp", ".hpp", ".rb", ".php", ".toml",
    ".gradle", ".dockerfile", ".service",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
ZIP_PACKAGE_EXTENSIONS = {".jar", ".war", ".ear", ".whl", ".apk"}
ARCHIVE_EXTENSIONS = {".zip", ".tar", ".tgz", ".tbz", ".tbz2", ".gz", ".bz2", ".xz", ".7z", ".rar"}
BINARY_PACKAGE_EXTENSIONS = {".rpm", ".deb", ".bin", ".run", ".dmg", ".pkg", ".msi", ".exe", ".iso"}
LEGACY_OFFICE_TARGETS = {".doc": ".docx", ".xls": ".xlsx", ".ppt": ".pptx"}
OFFICE_PREVIEW_EXTENSIONS = {
    ".doc", ".docx", ".xls", ".xlsx", ".xlsm", ".ppt", ".pptx",
}
def _office_environment(temp_root: Path) -> dict[str, str]:
    """Return a minimal environment without forwarding application secrets."""
    environment = {
        "HOME": str(temp_root),
        "TMPDIR": str(temp_root),
        "PATH": os.environ.get("PATH", "/usr/bin:/bin:/usr/sbin:/sbin"),
        "LANG": os.environ.get("LANG", "C.UTF-8"),
        "LC_ALL": os.environ.get("LC_ALL", os.environ.get("LANG", "C.UTF-8")),
        "SAL_DISABLE_OPENCL": "1",
    }
    if os.name == "nt" and os.environ.get("SYSTEMROOT"):
        environment["SYSTEMROOT"] = os.environ["SYSTEMROOT"]
    return environment


@dataclass
class ParsedAttachment:
    kind: str
    preview_type: str
    status: str
    text: str = ""
    error: str = ""
    metadata: dict = field(default_factory=dict)


@dataclass
class RenderedOfficePreview:
    status: str
    error: str = ""
    size: int = 0
    pages: int = 0
    converter: str = "libreoffice"


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self._ignored_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self._ignored_depth += 1
        elif tag.lower() in {"p", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self._ignored_depth:
            self._ignored_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self._ignored_depth:
            self.parts.append(data)


def _compound_extension(name: str) -> str:
    lower = name.lower()
    for suffix in (".tar.gz", ".tar.bz2", ".tar.xz"):
        if lower.endswith(suffix):
            return suffix
    return Path(lower).suffix


def _bounded(text: str, max_chars: int = MAX_EXTRACTED_CHARS) -> tuple[str, bool]:
    value = (text or "").replace("\x00", "").strip()
    if len(value) <= max_chars:
        return value, False
    head = value[: max_chars // 2]
    tail = value[-max_chars // 2 :]
    return f"{head}\n\n... [文件内容已截断] ...\n\n{tail}", True


def _read_bounded_bytes(path: Path, max_bytes: int = MAX_TEXT_BYTES) -> tuple[bytes, bool]:
    size = path.stat().st_size
    if size <= max_bytes:
        return path.read_bytes(), False
    half = max_bytes // 2
    with path.open("rb") as stream:
        head = stream.read(half)
        stream.seek(max(0, size - half))
        tail = stream.read(half)
    return head + b"\n\n... [binary gap] ...\n\n" + tail, True


def _decode_text(data: bytes) -> str:
    if not data:
        return ""
    if b"\x00" in data[:4096] and not (data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff")):
        raise ValueError("文件内容看起来是二进制，不能按文本解析")
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return data.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return data.decode("utf-8", errors="replace")


def _parse_plain_text(path: Path, extension: str) -> ParsedAttachment:
    data, byte_truncated = _read_bounded_bytes(path)
    text = _decode_text(data)
    if extension in {".html", ".htm"}:
        parser = _HTMLTextExtractor()
        parser.feed(text)
        text = "".join(parser.parts)
    text, char_truncated = _bounded(text)
    return ParsedAttachment(
        kind="text",
        preview_type="text",
        status="ready",
        text=text,
        metadata={
            "characters": len(text),
            "lines": text.count("\n") + (1 if text else 0),
            "truncated": byte_truncated or char_truncated,
        },
    )


def _parse_pdf(path: Path) -> ParsedAttachment:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ParsedAttachment(
            kind="pdf",
            preview_type="pdf",
            status="unsupported",
            error="缺少 pypdf，PDF 可以预览和下载，但暂时不能提取文字",
        )

    reader = PdfReader(str(path), strict=False)
    encrypted = bool(reader.is_encrypted)
    if encrypted:
        try:
            reader.decrypt("")
        except Exception:
            pass
    parts: list[str] = []
    page_count = len(reader.pages)
    parsed_pages = 0
    for index, page in enumerate(reader.pages[:300], start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:
            page_text = f"[第 {index} 页文字提取失败: {exc}]"
        if page_text.strip():
            parts.append(f"## 第 {index} 页\n{page_text.strip()}")
        parsed_pages = index
        if sum(len(part) for part in parts) >= MAX_EXTRACTED_CHARS:
            break
    text, truncated = _bounded("\n\n".join(parts))
    status = "ready" if text else "metadata_only"
    error = "" if text else "PDF 没有可提取文字，可能是扫描件或已加密"
    return ParsedAttachment(
        kind="pdf",
        preview_type="pdf",
        status=status,
        text=text,
        error=error,
        metadata={
            "pages": page_count,
            "parsed_pages": parsed_pages,
            "encrypted": encrypted,
            "truncated": truncated or parsed_pages < page_count,
        },
    )


def _read_zip_member(archive: zipfile.ZipFile, name: str) -> bytes:
    info = archive.getinfo(name)
    if info.file_size > MAX_XML_MEMBER_BYTES:
        raise ValueError(f"Office XML 成员过大: {name}")
    with archive.open(info) as stream:
        data = stream.read(MAX_XML_MEMBER_BYTES + 1)
    if len(data) > MAX_XML_MEMBER_BYTES:
        raise ValueError(f"Office XML 成员超过解析上限: {name}")
    return data


def _validate_office_archive(archive: zipfile.ZipFile) -> None:
    entries = archive.infolist()
    if len(entries) > MAX_OFFICE_ARCHIVE_ENTRIES:
        raise ValueError("Office 文件成员数量超过解析上限")
    total_uncompressed = sum(max(0, item.file_size) for item in entries)
    if total_uncompressed > MAX_OFFICE_TOTAL_UNCOMPRESSED_BYTES:
        raise ValueError("Office 文件解压后内容超过 256 MB 解析上限")


def _xml_text(data: bytes, text_tag_suffix: str = "}t") -> str:
    root = ElementTree.fromstring(data)
    parts: list[str] = []
    for element in root.iter():
        if element.tag.endswith(text_tag_suffix) and element.text:
            parts.append(element.text)
        elif element.tag.endswith("}p") or element.tag.endswith("}tr"):
            parts.append("\n")
    return " ".join(part for part in parts if part).replace(" \n ", "\n")


def _parse_docx(path: Path) -> ParsedAttachment:
    with zipfile.ZipFile(path) as archive:
        _validate_office_archive(archive)
        names = set(archive.namelist())
        if "word/document.xml" not in names:
            raise ValueError("DOCX 中缺少 word/document.xml")
        parts = [_xml_text(_read_zip_member(archive, "word/document.xml"))]
        for name in sorted(names):
            if re.fullmatch(r"word/(header|footer)\d+\.xml", name):
                parts.append(_xml_text(_read_zip_member(archive, name)))
    text, truncated = _bounded("\n\n".join(parts))
    return ParsedAttachment(
        kind="document",
        preview_type="text",
        status="ready" if text else "metadata_only",
        text=text,
        error="" if text else "文档中没有可提取文字",
        metadata={"characters": len(text), "truncated": truncated},
    )


def _natural_number(name: str) -> int:
    match = re.search(r"(\d+)(?=\.[^.]+$)", name)
    return int(match.group(1)) if match else 0


def _parse_pptx(path: Path) -> ParsedAttachment:
    parts: list[str] = []
    characters = 0
    parsed_slides = 0
    with zipfile.ZipFile(path) as archive:
        _validate_office_archive(archive)
        slides = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)),
            key=_natural_number,
        )
        for index, name in enumerate(slides, start=1):
            text = _xml_text(_read_zip_member(archive, name))
            parsed_slides = index
            if text.strip():
                section = f"## 第 {index} 页\n{text.strip()}"
                parts.append(section)
                characters += len(section)
            if characters >= MAX_EXTRACTED_CHARS:
                break
    text, truncated = _bounded("\n\n".join(parts))
    return ParsedAttachment(
        kind="presentation",
        preview_type="text",
        status="ready" if text else "metadata_only",
        text=text,
        error="" if text else "演示文稿中没有可提取文字",
        metadata={
            "slides": len(slides),
            "parsed_slides": parsed_slides,
            "characters": len(text),
            "truncated": truncated or parsed_slides < len(slides),
        },
    )


def _parse_xlsx(path: Path) -> ParsedAttachment:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ParsedAttachment(
            kind="spreadsheet",
            preview_type="none",
            status="unsupported",
            error="缺少 openpyxl，暂时不能解析 Excel 内容",
        )

    with zipfile.ZipFile(path) as archive:
        _validate_office_archive(archive)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    parts: list[str] = []
    total_rows = 0
    characters = 0
    try:
        for sheet in workbook.worksheets[:50]:
            heading = f"## 工作表: {sheet.title}"
            parts.append(heading)
            characters += len(heading)
            sheet_rows = 0
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if not any(values):
                    continue
                line = "\t".join(values).rstrip()
                parts.append(line)
                characters += len(line)
                sheet_rows += 1
                total_rows += 1
                if sheet_rows >= 5_000 or characters >= MAX_EXTRACTED_CHARS:
                    parts.append("... [工作表内容已截断] ...")
                    break
            if characters >= MAX_EXTRACTED_CHARS:
                break
    finally:
        workbook.close()
    text, truncated = _bounded("\n".join(parts))
    return ParsedAttachment(
        kind="spreadsheet",
        preview_type="text",
        status="ready" if text else "metadata_only",
        text=text,
        error="" if text else "表格中没有可提取内容",
        metadata={
            "sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames[:50],
            "parsed_rows": total_rows,
            "truncated": truncated,
        },
    )


def _parse_converted_office(path: Path, target_extension: str) -> ParsedAttachment:
    if target_extension == ".docx":
        return _parse_docx(path)
    if target_extension == ".xlsx":
        return _parse_xlsx(path)
    if target_extension == ".pptx":
        return _parse_pptx(path)
    raise ValueError(f"不支持的 Office 转换目标: {target_extension}")


def _terminate_process_tree(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return
    try:
        if os.name == "posix":
            os.killpg(process.pid, signal.SIGTERM)
        else:
            process.terminate()
    except (OSError, ProcessLookupError):
        pass
    try:
        process.communicate(timeout=2)
        return
    except subprocess.TimeoutExpired:
        pass
    try:
        if os.name == "posix":
            os.killpg(process.pid, signal.SIGKILL)
        else:
            process.kill()
    except (OSError, ProcessLookupError):
        pass
    process.communicate()


def _run_office_conversion(command: list[str], *, environment: dict[str, str]) -> subprocess.CompletedProcess:
    """Run one LibreOffice process with a timeout and process-tree cleanup."""
    process = subprocess.Popen(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=environment,
        cwd=environment.get("TMPDIR") or None,
        start_new_session=os.name == "posix",
    )
    try:
        stdout, stderr = process.communicate(timeout=OFFICE_CONVERSION_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired:
        _terminate_process_tree(process)
        raise
    return subprocess.CompletedProcess(command, process.returncode, stdout, stderr)


def _office_preview_source_error(path: Path, original_name: str) -> str:
    """Reject oversized or structurally unsafe Office input before conversion."""
    extension = _compound_extension(original_name)
    try:
        size = path.stat().st_size
    except OSError as exc:
        return f"Office 文件无法读取: {exc}"
    if size <= 0:
        return "Office 文件为空，无法生成版式预览"
    if size > MAX_LEGACY_OFFICE_BYTES:
        return "Office 文件超过 100 MB 预览上限"
    if extension in {".docx", ".xlsx", ".xlsm", ".pptx"}:
        try:
            with zipfile.ZipFile(path) as archive:
                _validate_office_archive(archive)
        except (OSError, ValueError, zipfile.BadZipFile) as exc:
            return f"Office 文件结构不安全或已损坏: {exc}"
    return ""


def render_office_pdf(
    path: Path,
    original_name: str,
    destination: Path,
) -> RenderedOfficePreview:
    """Render an Office document to an atomic PDF sidecar for layout preview."""
    extension = _compound_extension(original_name)
    if extension not in OFFICE_PREVIEW_EXTENSIONS:
        return RenderedOfficePreview(status="unsupported", error="该格式不支持 Office 版式预览")
    source_error = _office_preview_source_error(path, original_name)
    if source_error:
        return RenderedOfficePreview(status="unsupported", error=source_error)

    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        return RenderedOfficePreview(
            status="unsupported",
            error="当前环境未安装 LibreOffice，无法生成 Office 版式预览",
        )

    with tempfile.TemporaryDirectory(prefix="opsane-office-preview-") as temp_value:
        temp_root = Path(temp_value)
        output_dir = temp_root / "output"
        profile_dir = temp_root / "profile"
        output_dir.mkdir()
        profile_dir.mkdir()
        command = [
            executable,
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--safe-mode",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(path),
        ]
        environment = _office_environment(temp_root)
        try:
            result = _run_office_conversion(command, environment=environment)
        except subprocess.TimeoutExpired:
            return RenderedOfficePreview(
                status="error",
                error=f"Office 版式转换超过 {OFFICE_CONVERSION_TIMEOUT_SECONDS} 秒已停止",
            )
        except OSError as exc:
            return RenderedOfficePreview(status="error", error=f"Office 预览转换器启动失败: {exc}")

        candidates = sorted(output_dir.glob("*.pdf"))
        if result.returncode != 0 or len(candidates) != 1:
            detail = result.stderr.decode("utf-8", errors="replace").strip()
            if not detail:
                detail = result.stdout.decode("utf-8", errors="replace").strip()
            return RenderedOfficePreview(
                status="error",
                error=f"Office 版式转换失败: {detail[:400] or '没有生成 PDF'}",
            )

        rendered = candidates[0]
        if rendered.is_symlink() or not rendered.is_file():
            return RenderedOfficePreview(status="error", error="Office 版式转换结果路径异常")
        rendered_size = rendered.stat().st_size
        if rendered_size <= 0 or rendered_size > MAX_OFFICE_PREVIEW_BYTES:
            return RenderedOfficePreview(status="error", error="Office 版式预览 PDF 大小异常")
        with rendered.open("rb") as stream:
            if stream.read(5) != b"%PDF-":
                return RenderedOfficePreview(status="error", error="Office 版式转换结果不是有效 PDF")

        try:
            from pypdf import PdfReader

            pages = len(PdfReader(str(rendered), strict=False).pages)
        except Exception as exc:
            return RenderedOfficePreview(status="error", error=f"Office 版式转换结果无法解析: {exc}")
        if pages <= 0:
            return RenderedOfficePreview(status="error", error="Office 版式转换结果没有可预览页面")

        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix=f".{destination.name}.",
                suffix=".tmp",
                dir=destination.parent,
                delete=False,
            ) as temporary:
                temporary_path = Path(temporary.name)
                with rendered.open("rb") as source:
                    shutil.copyfileobj(source, temporary)
                temporary.flush()
                os.fsync(temporary.fileno())
            os.replace(temporary_path, destination)
        except OSError as exc:
            return RenderedOfficePreview(status="error", error=f"Office 版式预览保存失败: {exc}")
        finally:
            if temporary_path:
                temporary_path.unlink(missing_ok=True)
        return RenderedOfficePreview(
            status="ready",
            size=rendered_size,
            pages=pages,
        )


def _parse_legacy_office(path: Path, extension: str) -> ParsedAttachment:
    """Convert one legacy Office file with an isolated temporary profile."""
    target_extension = LEGACY_OFFICE_TARGETS[extension]
    kind = {
        ".doc": "document",
        ".xls": "spreadsheet",
        ".ppt": "presentation",
    }[extension]
    if path.stat().st_size > MAX_LEGACY_OFFICE_BYTES:
        return ParsedAttachment(
            kind=kind,
            preview_type="none",
            status="unsupported",
            error="旧版 Office 文件超过 100 MB 安全转换上限",
            metadata={"source_format": extension.lstrip(".")},
        )

    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        return ParsedAttachment(
            kind=kind,
            preview_type="none",
            status="unsupported",
            error="当前环境未安装 LibreOffice，无法安全解析旧版 Office 文件",
            metadata={"source_format": extension.lstrip(".")},
        )

    with tempfile.TemporaryDirectory(prefix="opsane-office-") as temp_value:
        temp_root = Path(temp_value)
        output_dir = temp_root / "output"
        profile_dir = temp_root / "profile"
        output_dir.mkdir()
        profile_dir.mkdir()
        command = [
            executable,
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--safe-mode",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--convert-to",
            target_extension.lstrip("."),
            "--outdir",
            str(output_dir),
            str(path),
        ]
        environment = _office_environment(temp_root)
        try:
            result = _run_office_conversion(command, environment=environment)
        except subprocess.TimeoutExpired:
            return ParsedAttachment(
                kind=kind,
                preview_type="none",
                status="error",
                error=f"旧版 Office 文件转换超过 {OFFICE_CONVERSION_TIMEOUT_SECONDS} 秒已停止",
                metadata={"source_format": extension.lstrip(".")},
            )
        except OSError as exc:
            return ParsedAttachment(
                kind=kind,
                preview_type="none",
                status="error",
                error=f"旧版 Office 转换器启动失败: {exc}",
                metadata={"source_format": extension.lstrip(".")},
            )

        candidates = sorted(output_dir.glob(f"*{target_extension}"))
        if result.returncode != 0 or len(candidates) != 1:
            detail = result.stderr.decode("utf-8", errors="replace").strip()
            if not detail:
                detail = result.stdout.decode("utf-8", errors="replace").strip()
            return ParsedAttachment(
                kind=kind,
                preview_type="none",
                status="error",
                error=f"旧版 Office 文件转换失败: {detail[:400] or '没有生成转换文件'}",
                metadata={"source_format": extension.lstrip(".")},
            )

        converted_path = candidates[0]
        if converted_path.stat().st_size > MAX_LEGACY_OFFICE_BYTES * 2:
            return ParsedAttachment(
                kind=kind,
                preview_type="none",
                status="error",
                error="旧版 Office 转换结果异常过大，已停止解析",
                metadata={"source_format": extension.lstrip(".")},
            )
        try:
            parsed = _parse_converted_office(converted_path, target_extension)
        except Exception as exc:
            return ParsedAttachment(
                kind=kind,
                preview_type="none",
                status="error",
                error=f"旧版 Office 转换结果解析失败: {type(exc).__name__}: {exc}",
                metadata={
                    "source_format": extension.lstrip("."),
                    "converted_format": target_extension.lstrip("."),
                    "converter": "libreoffice",
                },
            )
        parsed.metadata.update(
            {
                "source_format": extension.lstrip("."),
                "converted_format": target_extension.lstrip("."),
                "converter": "libreoffice",
            }
        )
        return parsed


@lru_cache(maxsize=4)
def _tesseract_languages(executable: str) -> tuple[str, ...]:
    try:
        result = subprocess.run(
            [executable, "--list-langs"],
            capture_output=True,
            check=False,
            timeout=10,
        )
    except (OSError, subprocess.SubprocessError):
        return ()
    output = result.stdout.decode("utf-8", errors="replace")
    return tuple(
        line.strip()
        for line in output.splitlines()
        if line.strip() and not line.lower().startswith("list of available")
    )


def _preferred_ocr_languages(available: tuple[str, ...]) -> tuple[str, ...]:
    values = set(available)
    preferred = tuple(
        language
        for language in ("chi_sim", "chi_tra", "eng")
        if language in values
    )
    if preferred:
        return preferred
    return (available[0],) if available else ()


@lru_cache(maxsize=1)
def _macos_vision_binary() -> str:
    """Compile the bundled local Vision OCR helper into a content-addressed cache."""
    if sys.platform != "darwin":
        return ""
    swiftc = shutil.which("swiftc")
    source = Path(__file__).with_name("macos_vision_ocr.swift")
    if not swiftc or not source.is_file():
        return ""
    try:
        source_bytes = source.read_bytes()
        cache_dir = Path(tempfile.gettempdir()) / "opsane-vision-ocr"
        cache_dir.mkdir(parents=True, exist_ok=True)
        binary = cache_dir / f"vision-ocr-{sha256(source_bytes).hexdigest()[:16]}"
        if not binary.is_file():
            module_cache = cache_dir / "module-cache"
            module_cache.mkdir(parents=True, exist_ok=True)
            environment = os.environ.copy()
            environment["CLANG_MODULE_CACHE_PATH"] = str(module_cache)
            environment["SWIFT_MODULECACHE_PATH"] = str(module_cache)
            result = subprocess.run(
                [
                    swiftc,
                    "-O",
                    "-module-cache-path",
                    str(module_cache),
                    str(source),
                    "-o",
                    str(binary),
                ],
                capture_output=True,
                check=False,
                env=environment,
                timeout=90,
            )
            if result.returncode != 0:
                return ""
            binary.chmod(0o700)
        return str(binary)
    except (OSError, subprocess.SubprocessError):
        return ""


def _parse_image_with_macos_vision(path: Path, metadata: dict) -> ParsedAttachment | None:
    executable = _macos_vision_binary()
    if not executable:
        return None
    try:
        result = subprocess.run(
            [executable, str(path)],
            capture_output=True,
            check=False,
            timeout=OCR_TIMEOUT_SECONDS,
        )
    except (OSError, subprocess.SubprocessError):
        return None
    if result.returncode != 0:
        return None
    raw_text = result.stdout.decode("utf-8", errors="replace")
    text, truncated = _bounded(raw_text)
    if not text:
        return None
    metadata.update(
        {
            "analysis_supported": True,
            "ocr_status": "ready",
            "ocr_engine": "apple_vision",
            "ocr_languages": ["zh-Hans", "zh-Hant", "en-US"],
            "characters": len(text),
            "truncated": truncated,
        }
    )
    return ParsedAttachment(
        kind="image",
        preview_type="image",
        status="ready",
        text=f"## 图片文字识别（本地 Vision OCR）\n{text}",
        metadata=metadata,
    )


def _parse_image(path: Path) -> ParsedAttachment:
    """Preview an image and extract visible text without executing its content."""
    metadata = {
        "analysis_supported": False,
        "ocr_status": "pending",
        "ocr_engine": "tesseract",
    }
    size = path.stat().st_size
    if size > MAX_OCR_IMAGE_BYTES:
        metadata["ocr_status"] = "too_large"
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error="图片可以预览，但超过 30 MB OCR 安全上限",
            metadata=metadata,
        )

    vision_result = _parse_image_with_macos_vision(path, metadata)
    if vision_result is not None:
        return vision_result

    executable = shutil.which("tesseract")
    if not executable:
        metadata["ocr_status"] = "unavailable"
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error="图片可以预览，但当前环境未安装 Tesseract OCR",
            metadata=metadata,
        )

    languages = _preferred_ocr_languages(_tesseract_languages(executable))
    metadata["ocr_languages"] = list(languages)
    if not languages:
        metadata["ocr_status"] = "unavailable"
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error="图片可以预览，但 Tesseract 没有可用的语言数据",
            metadata=metadata,
        )

    try:
        result = subprocess.run(
            [
                executable,
                str(path),
                "stdout",
                "-l",
                "+".join(languages),
                "--psm",
                "3",
            ],
            capture_output=True,
            check=False,
            timeout=OCR_TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired:
        metadata["ocr_status"] = "timeout"
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error=f"图片可以预览，但 OCR 超过 {OCR_TIMEOUT_SECONDS} 秒已停止",
            metadata=metadata,
        )
    except OSError as exc:
        metadata["ocr_status"] = "error"
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error=f"图片可以预览，但 OCR 启动失败: {exc}",
            metadata=metadata,
        )

    if result.returncode != 0:
        metadata["ocr_status"] = "error"
        error = result.stderr.decode("utf-8", errors="replace").strip()
        return ParsedAttachment(
            kind="image",
            preview_type="image",
            status="metadata_only",
            error=f"图片可以预览，但 OCR 识别失败: {error[:300] or '未知错误'}",
            metadata=metadata,
        )

    raw_text = result.stdout.decode("utf-8", errors="replace")
    text, truncated = _bounded(raw_text)
    metadata.update(
        {
            "analysis_supported": bool(text),
            "ocr_status": "ready" if text else "empty",
            "characters": len(text),
            "truncated": truncated,
        }
    )
    return ParsedAttachment(
        kind="image",
        preview_type="image",
        status="ready" if text else "metadata_only",
        text=f"## 图片文字识别（OCR）\n{text}" if text else "",
        error="" if text else "图片可以预览，但没有识别出可用文字",
        metadata=metadata,
    )


def _archive_listing(path: Path, extension: str) -> tuple[list[str], dict]:
    entries: list[str] = []
    total_entries = 0
    total_uncompressed = 0
    if zipfile.is_zipfile(path):
        with zipfile.ZipFile(path) as archive:
            for info in archive.infolist():
                total_entries += 1
                total_uncompressed += max(0, info.file_size)
                if len(entries) < MAX_ARCHIVE_ENTRIES:
                    marker = "/" if info.is_dir() else ""
                    entries.append(f"{info.filename}{marker}\t{info.file_size} bytes")
    elif tarfile.is_tarfile(path):
        with tarfile.open(path, mode="r:*") as archive:
            for member in archive:
                total_entries += 1
                total_uncompressed += max(0, member.size)
                if len(entries) < MAX_ARCHIVE_ENTRIES:
                    marker = "/" if member.isdir() else ""
                    entries.append(f"{member.name}{marker}\t{member.size} bytes")
                if total_entries >= MAX_ARCHIVE_ENTRIES * 10:
                    break
    else:
        return [], {"format": extension.lstrip("."), "entries": 0, "listing_supported": False}
    return entries, {
        "format": extension.lstrip("."),
        "entries": total_entries,
        "listed_entries": len(entries),
        "total_uncompressed_size": total_uncompressed,
        "listing_truncated": total_entries > len(entries),
        "listing_supported": True,
    }


def _parse_archive_or_package(path: Path, extension: str, package: bool) -> ParsedAttachment:
    entries, metadata = _archive_listing(path, extension)
    label = "安装包/制品" if package else "压缩包"
    lines = [f"# {label}元数据", f"- 格式: {extension or 'unknown'}", f"- 文件大小: {path.stat().st_size} bytes"]
    if entries:
        lines.extend(["", "## 文件清单（未解压、未执行）", *entries])
    elif not metadata.get("listing_supported"):
        lines.extend(["", "该格式仅记录元数据，不读取或执行包内内容。"])
    text, truncated = _bounded("\n".join(lines))
    metadata["truncated"] = truncated or bool(metadata.get("listing_truncated"))
    return ParsedAttachment(
        kind="package" if package else "archive",
        preview_type="text",
        status="metadata_only",
        text=text,
        metadata=metadata,
    )


def _magic_prefix(path: Path) -> str:
    with path.open("rb") as stream:
        return stream.read(16).hex(" ")


def parse_attachment(path: Path, original_name: str, media_type: str = "") -> ParsedAttachment:
    """Parse one uploaded file without executing or extracting it."""
    extension = _compound_extension(original_name)
    try:
        if extension == ".pdf" or media_type == "application/pdf":
            return _parse_pdf(path)
        if extension in LEGACY_OFFICE_TARGETS:
            return _parse_legacy_office(path, extension)
        if extension == ".docx":
            return _parse_docx(path)
        if extension == ".pptx":
            return _parse_pptx(path)
        if extension in {".xlsx", ".xlsm"}:
            return _parse_xlsx(path)
        if extension in IMAGE_EXTENSIONS and media_type != "image/svg+xml":
            return _parse_image(path)
        if extension in ZIP_PACKAGE_EXTENSIONS:
            return _parse_archive_or_package(path, extension, package=True)
        if extension in BINARY_PACKAGE_EXTENSIONS:
            parsed = _parse_archive_or_package(path, extension, package=True)
            parsed.metadata["magic_prefix"] = _magic_prefix(path)
            return parsed
        if extension in ARCHIVE_EXTENSIONS or extension.startswith(".tar."):
            return _parse_archive_or_package(path, extension, package=False)
        if extension in TEXT_EXTENSIONS or media_type.startswith("text/"):
            return _parse_plain_text(path, extension)

        # A bounded text probe lets extensionless config files remain useful.
        data, _ = _read_bounded_bytes(path, max_bytes=min(MAX_TEXT_BYTES, 256 * 1024))
        try:
            text = _decode_text(data)
        except ValueError:
            text = ""
        printable = sum(character.isprintable() or character in "\r\n\t" for character in text)
        if text and printable / max(1, len(text)) > 0.85:
            return _parse_plain_text(path, extension)
        return ParsedAttachment(
            kind="other",
            preview_type="none",
            status="metadata_only",
            error="该二进制格式不支持内容解析，可下载原文件",
            metadata={"magic_prefix": _magic_prefix(path)},
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile, tarfile.TarError, ElementTree.ParseError) as exc:
        return ParsedAttachment(
            kind="other",
            preview_type="none",
            status="error",
            error=f"文件解析失败: {exc}",
            metadata={"magic_prefix": _magic_prefix(path) if path.exists() else ""},
        )
    except Exception as exc:
        return ParsedAttachment(
            kind="other",
            preview_type="none",
            status="error",
            error=f"文件解析失败: {type(exc).__name__}: {exc}",
        )
